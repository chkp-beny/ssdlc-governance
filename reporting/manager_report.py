from .product_report import ProductReport
from .app_status_report import AppStatusReport
from CONSTANTS import PILLAR_PRODUCTS, PRODUCT_SCM_TYPE, PRODUCT_SCM_INSTANCE


class ManagerReport(ProductReport):
    report_type = "manager"
    columns = [
        'product_pillar', 'product', 'scm', 'repo_name',
        'repo_owner', 'repo_owner_title', 'group_manager', 'director', 'vp',
        'status_scan_dependencies_jfrog', 'status_scan_sast_sonar',
        'critical_dependencies_vulnerabilities_jfrog',
        'high_dependencies_vulnerabilities_jfrog',
        'critical_code_vulnerabilities_sonar',
        'critical_code_secrets_sonar',
        'prevent_on_code_push_to_scm',
        'prevent_on_artifact_push_to_registry',
        'notes'
    ]
    output_format = "xlsx"

    def extract_repo_data(self, repo, product_name: str) -> dict:
        data = {col: "unhandled yet" for col in self.columns}
        data['product'] = product_name
        data['scm'] = PRODUCT_SCM_INSTANCE.get(product_name, '')
        # Pillar
        product_pillar = 'unhandled yet'
        for pillar, products in PILLAR_PRODUCTS.items():
            if product_name in products:
                product_pillar = pillar
                break
        data['product_pillar'] = product_pillar
        data['repo_name'] = getattr(repo, 'get_repository_name', lambda: 'unknown')()
        data['repo_owner'] = repo.get_primary_owner_email() if hasattr(repo, 'get_primary_owner_email') else "unknown"
        data['repo_owner_title'] = repo.get_primary_owner_title() if hasattr(repo, 'get_primary_owner_title') else "unknown"
        data['group_manager'] = repo.get_primary_owner_general_manager() if hasattr(repo, 'get_primary_owner_general_manager') else "unknown"
        data['director'] = repo.get_primary_owner_director() if hasattr(repo, 'get_primary_owner_director') else "unknown"
        data['vp'] = repo.get_primary_owner_vp() if hasattr(repo, 'get_primary_owner_vp') else "unknown"
        # JFrog/CI status
        data['status_scan_dependencies_jfrog'] = False
        data['status_scan_sast_sonar'] = False
        # Vulnerabilities
        data['critical_dependencies_vulnerabilities_jfrog'] = 0
        data['high_dependencies_vulnerabilities_jfrog'] = 0
        data['critical_code_vulnerabilities_sonar'] = 0
        data['critical_code_secrets_sonar'] = 0
        data['prevent_on_code_push_to_scm'] = "unhandled yet"
        data['prevent_on_artifact_push_to_registry'] = "unhandled yet"

        if hasattr(repo, 'ci_status') and repo.ci_status:
            if hasattr(repo.ci_status, 'jfrog_status') and repo.ci_status.jfrog_status:
                data['status_scan_dependencies_jfrog'] = repo.ci_status.jfrog_status.is_exist
            if hasattr(repo.ci_status, 'sonar_status') and repo.ci_status.sonar_status:
                data['status_scan_sast_sonar'] = repo.ci_status.sonar_status.is_exist

        # Handle JFrog integration status and vulnerabilities based on the status we just set
        if not data['status_scan_dependencies_jfrog']:
            # JFrog not integrated - set "Not Integrated"
            data['critical_dependencies_vulnerabilities_jfrog'] = "Not Integrated"
            data['high_dependencies_vulnerabilities_jfrog'] = "Not Integrated"
        elif hasattr(repo, 'vulnerabilities') and repo.vulnerabilities and hasattr(repo.vulnerabilities, 'dependencies_vulns') and repo.vulnerabilities.dependencies_vulns:
            # JFrog integrated and vulnerabilities exist - set actual counts
            deps_vuln = repo.vulnerabilities.dependencies_vulns
            data['critical_dependencies_vulnerabilities_jfrog'] = deps_vuln.critical_count
            data['high_dependencies_vulnerabilities_jfrog'] = deps_vuln.high_count
        # If JFrog integrated but no vulnerabilities, keep default 0 values

        # Handle SonarQube integration status and vulnerabilities based on the status we just set
        if not data['status_scan_sast_sonar']:
            # SonarQube not integrated - set "Not Integrated"
            data['critical_code_secrets_sonar'] = "Not Integrated"
            data['critical_code_vulnerabilities_sonar'] = "Not Integrated"
        elif hasattr(repo, 'vulnerabilities') and repo.vulnerabilities and hasattr(repo.vulnerabilities, 'code_issues') and repo.vulnerabilities.code_issues:
            # SonarQube integrated and vulnerabilities exist - set actual counts
            code_issues = repo.vulnerabilities.code_issues
            data['critical_code_secrets_sonar'] = code_issues.get_secrets_count() if hasattr(code_issues, 'get_secrets_count') else 0
            data['critical_code_vulnerabilities_sonar'] = code_issues.get_critical_vulnerability_count() if hasattr(code_issues, 'get_critical_vulnerability_count') else 0
        # If SonarQube integrated but no vulnerabilities, keep default 0 values
        data['notes'] = repo.get_notes_display()
        return data

    def generate_report(self, output_dir: str) -> str:
        """
        Generate manager report with both DATA and App Status sheets.
        """
        from datetime import datetime
        import os
        
        # Collect all repository data
        all_rows = []
        for product in self.load_all_products():
            for repo in product.repos:
                all_rows.append(self.extract_repo_data(repo, product.name))
        
        if not all_rows:
            print("No data to report.")
            return ""
        
        if self.output_format == "xlsx":
            # Excel with multiple sheets
            xlsx_filename = f"{self.report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            xlsx_path = os.path.join(output_dir, xlsx_filename)
            
            try:
                import pandas as pd
                from openpyxl import Workbook
                
                # Create workbook
                workbook = Workbook()
                
                # Remove default sheet
                workbook.remove(workbook.active)
                
                # Create DATA sheet (detailed report)
                df = pd.DataFrame(all_rows)
                df = df[self.columns]
                
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='DATA', index=False)
                
                # Reopen to add App Status sheet
                from openpyxl import load_workbook
                workbook = load_workbook(xlsx_path)
                
                # Create App Status sheet
                app_status_report = AppStatusReport(all_rows)
                app_status_report.export_to_excel(workbook, "App Status")
                
                # Save the workbook with both sheets
                workbook.save(xlsx_path)
                
                print(f"✅ Generated XLSX file with DATA and App Status sheets: {xlsx_path}")
                return xlsx_path
                
            except ImportError as e:
                print(f"❌ Missing required dependencies for XLSX generation: {e}")
                return ""
            except Exception as e:
                print(f"❌ Error generating XLSX file: {e}")
                return ""
        else:
            # Fall back to parent implementation for other formats
            return super().generate_report(output_dir)
